"""
Huliot DU Calculation Web App
Processes DU (Drainage Unit) calculation sheets per EN 12056-2
Deployable on Streamlit Cloud – no local installation needed
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import io
import math
from datetime import datetime
import copy

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Huliot DU Calculator",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS (mobile-friendly) ──────────────────────────────────────────────
st.markdown("""
<style>
    /* Compact header */
    .main-title {
        font-size: 1.6rem; font-weight: 700;
        color: #1a3a5c; margin-bottom: 0;
    }
    .sub-title {
        font-size: 0.85rem; color: #555; margin-top: 2px;
    }
    /* Metric cards */
    .metric-card {
        background: #f0f6ff; border-left: 4px solid #1a6fc4;
        border-radius: 6px; padding: 10px 14px; margin: 6px 0;
    }
    .metric-label { font-size: 0.75rem; color: #555; font-weight: 600; }
    .metric-value { font-size: 1.4rem; font-weight: 700; color: #1a3a5c; }
    /* Section heading */
    .section-head {
        background: #1a3a5c; color: white;
        padding: 6px 12px; border-radius: 4px;
        font-size: 0.9rem; font-weight: 600; margin: 14px 0 8px;
    }
    /* Responsive table */
    [data-testid="stDataFrame"] { width: 100% !important; }
    /* Download button */
    [data-testid="stDownloadButton"] > button {
        background: #1a6fc4; color: white; border: none;
        border-radius: 6px; padding: 0.5rem 1.2rem;
        font-weight: 600; width: 100%;
    }
    /* Sidebar */
    section[data-testid="stSidebar"] { background: #f8faff; }
    /* Warning banner */
    .warn-box {
        background: #fff8e1; border-left: 4px solid #f59f00;
        padding: 8px 12px; border-radius: 4px; font-size: 0.82rem;
    }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# ── EN 12056-2 CALCULATION ENGINE ────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

# Flow coefficients (K factor) per EN 12056-2 System III (India standard)
K_FACTOR = 0.7   # For mixed use / intermittent systems

# Minimum pipe sizes by flow (mm) for gravity drainage
PIPE_SIZES = [
    (0.0,  0.5,  50,  1.2),
    (0.5,  0.9,  75,  2.5),
    (0.9,  1.5, 100,  3.5),
    (1.5,  2.9, 110,  4.5),   # Huliot HT Pro 110
    (2.9,  5.2, 110,  5.2),   # Huliot HT Pro 110 with swept tee
    (5.2,  7.6, 125,  7.6),   # Huliot HT Pro 125 with swept tee
    (7.6, 12.4, 160, 12.4),
    (12.4, 999, 200, 24.0),
]

def du_to_flow(du: float) -> float:
    """Convert Drainage Units → flow (l/s) per EN 12056-2."""
    if du <= 0:
        return 0.0
    return round(K_FACTOR * math.sqrt(du), 3)

def flow_to_pipe(flow: float) -> tuple:
    """Return (pipe_dia_mm, max_flow_l_s) for given flow."""
    for lo, hi, dia, cap in PIPE_SIZES:
        if lo <= flow <= hi:
            return dia, cap
    return 200, 24.0  # fallback

# Default fixture DU values (EN 12056-2 Annex A)
FIXTURE_DU = {
    "KITCHEN SINK": 0.8,
    "WM (WASHING MACHINE UP TO 6 KG CONSIDER)": 0.8,
    "DISH WASHER": 0.8,
    "W C- WITH 6 L CISTERN": 2.0,
    "FLOOR TRAP- Shower without plug consider": 0.6,
    "WASHBASIN": 0.5,
    "Trap": 1.5,
    "URINAL": 0.5,
    "BATHTUB": 1.5,
    "BIDET": 0.5,
}


# ══════════════════════════════════════════════════════════════════════════════
# ── FILE PARSING ─────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

def parse_du_sheet(file_bytes: bytes) -> dict:
    """
    Parse the uploaded DU Calculation Sheet.
    Returns dict: {section_name: DataFrame}
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    sections = {}
    current_section = None
    current_rows = []
    header_found = False

    for row in ws.iter_rows(values_only=True):
        # Detect section header (bold, merged, non-empty first cell)
        if row[0] is not None and all(c is None for c in row[1:4]):
            # Save previous section
            if current_section and current_rows:
                sections[current_section] = pd.DataFrame(
                    current_rows,
                    columns=["SR NO.", "FIXTURES", "QTY", "DU",
                             "FOR TOILET", "FLOORS", "TOTAL DU",
                             "TOTAL FLOW (l/s)", "REQ DIA (mm)"]
                )
            current_section = str(row[0]).strip()
            current_rows = []
            header_found = False
            continue

        # Detect column header row
        if row[1] == "FIXTURES":
            header_found = True
            continue

        # Data rows
        if header_found and row[0] is not None:
            try:
                sr = int(row[0])
                fixture = str(row[1]).strip() if row[1] else ""
                qty = float(row[2]) if row[2] not in (None, "") else 0.0
                du_val = float(row[3]) if row[3] not in (None, "") else 0.0
                for_toilet = int(row[4]) if row[4] not in (None, "") else 0
                floors = float(row[5]) if row[5] not in (None, "") else 0.0
                total_du = float(row[6]) if row[6] not in (None, "") else 0.0
                total_flow = float(row[7]) if row[7] not in (None, "") else 0.0
                req_dia = str(row[8]) if row[8] else ""
                current_rows.append([sr, fixture, qty, du_val, for_toilet,
                                      floors, total_du, total_flow, req_dia])
            except (ValueError, TypeError):
                continue

    # Save last section
    if current_section and current_rows:
        sections[current_section] = pd.DataFrame(
            current_rows,
            columns=["SR NO.", "FIXTURES", "QTY", "DU",
                     "FOR TOILET", "FLOORS", "TOTAL DU",
                     "TOTAL FLOW (l/s)", "REQ DIA (mm)"]
        )

    return sections


def recalculate(df: pd.DataFrame) -> pd.DataFrame:
    """Recompute TOTAL DU, TOTAL FLOW, REQ DIA for each fixture row."""
    df = df.copy()
    for i, row in df.iterrows():
        qty = float(row["QTY"]) if pd.notna(row["QTY"]) else 0
        du = float(row["DU"]) if pd.notna(row["DU"]) else 0
        floors = float(row["FLOORS"]) if pd.notna(row["FLOORS"]) else 0
        total_du = qty * du * floors
        df.at[i, "TOTAL DU"] = round(total_du, 2)
        flow = du_to_flow(total_du)
        df.at[i, "TOTAL FLOW (l/s)"] = flow
        if total_du > 0:
            dia, cap = flow_to_pipe(flow)
            df.at[i, "REQ DIA (mm)"] = f"{dia} mm ({cap} l/s cap)"
        else:
            df.at[i, "REQ DIA (mm)"] = "—"
    return df


# ══════════════════════════════════════════════════════════════════════════════
# ── EXCEL OUTPUT BUILDER ─────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

HULIOT_BLUE  = "1A3A5C"
HULIOT_LBLUE = "1A6FC4"
HEADER_BG    = "D9E8F5"
ALT_ROW_BG   = "F0F6FF"
WHITE        = "FFFFFF"
ORANGE       = "F59F00"
GREEN        = "2D7D2D"

def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _hdr_style(ws, row, col_count, title, bg=HULIOT_BLUE):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=col_count)

def _col_hdr(ws, row, headers, bg=HEADER_BG):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True, color=HULIOT_BLUE, size=9)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _border()

def build_excel(sections: dict, project_info: dict) -> bytes:
    """Build a formatted Excel workbook from processed sections."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ── Sheet 1: DU Summary ──────────────────────────────────────────────────
    ws = wb.create_sheet("DU Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 9
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 22

    # Title block
    ws.row_dimensions[1].height = 6
    ws.merge_cells("A2:I2")
    title_cell = ws["A2"]
    title_cell.value = "DRAINAGE UNIT (DU) CALCULATION SHEET"
    title_cell.font = Font(bold=True, color=WHITE, size=13)
    title_cell.fill = PatternFill("solid", fgColor=HULIOT_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28

    ws.merge_cells("A3:I3")
    sub = ws["A3"]
    sub.value = f"Project: {project_info.get('project','—')}  |  Prepared by: {project_info.get('engineer','—')}  |  Date: {project_info.get('date','—')}  |  Standard: EN 12056-2"
    sub.font = Font(color=WHITE, size=8)
    sub.fill = PatternFill("solid", fgColor=HULIOT_LBLUE)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18

    cur_row = 5
    COLS = ["SR NO.", "FIXTURES", "QTY", "DU",
            "FOR TOILET", "FLOORS", "TOTAL DU",
            "TOTAL FLOW (l/s)", "REQ DIA (mm)"]

    section_totals = []   # for summary tab

    for sec_name, df in sections.items():
        # Section header
        _hdr_style(ws, cur_row, 9, f"  ▸  {sec_name.upper()}")
        ws.row_dimensions[cur_row].height = 20
        cur_row += 1

        _col_hdr(ws, cur_row, COLS)
        ws.row_dimensions[cur_row].height = 26
        cur_row += 1

        total_du_sec = 0
        for idx, (_, row) in enumerate(df.iterrows()):
            bg = ALT_ROW_BG if idx % 2 == 0 else WHITE
            vals = [
                row["SR NO."], row["FIXTURES"],
                row["QTY"],    row["DU"],
                row["FOR TOILET"], row["FLOORS"],
                row["TOTAL DU"], row["TOTAL FLOW (l/s)"],
                row["REQ DIA (mm)"]
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=cur_row, column=c, value=v)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.border = _border()
                cell.font = Font(size=9)
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center" if c != 2 else "left",
                    wrap_text=(c == 2),
                )
                # Highlight DU column
                if c == 7 and isinstance(v, (int, float)) and v > 0:
                    cell.font = Font(bold=True, color=HULIOT_LBLUE, size=9)
            ws.row_dimensions[cur_row].height = 20
            total_du_sec += row["TOTAL DU"] if pd.notna(row["TOTAL DU"]) else 0
            cur_row += 1

        # Section totals row
        sec_flow = du_to_flow(total_du_sec)
        sec_dia, sec_cap = flow_to_pipe(sec_flow)
        total_vals = ["", "SECTION TOTAL", "", "", "", "",
                      round(total_du_sec, 2),
                      sec_flow,
                      f"{sec_dia} mm (cap {sec_cap} l/s)"]
        for c, v in enumerate(total_vals, 1):
            cell = ws.cell(row=cur_row, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor="FFF3CD")
            cell.font = Font(bold=True, size=9, color="7B4F00")
            cell.border = _border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[cur_row].height = 18
        cur_row += 2

        section_totals.append({
            "Section": sec_name,
            "Total DU": round(total_du_sec, 2),
            "Flow (l/s)": sec_flow,
            "Req. Pipe Dia (mm)": f"{sec_dia} mm",
        })

    # ── Sheet 2: Project Summary Report ──────────────────────────────────────
    ws2 = wb.create_sheet("Project Report")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 24

    ws2.merge_cells("A1:D1")
    h = ws2["A1"]
    h.value = "PROJECT DRAINAGE REPORT — HULIOT SYSTEMS"
    h.font = Font(bold=True, color=WHITE, size=12)
    h.fill = PatternFill("solid", fgColor=HULIOT_BLUE)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    info_rows = [
        ("Project Name", project_info.get("project", "—")),
        ("Engineer", project_info.get("engineer", "—")),
        ("Date", project_info.get("date", "—")),
        ("Standard", "EN 12056-2 (System III)"),
        ("K-Factor Used", str(K_FACTOR)),
    ]
    for r, (lbl, val) in enumerate(info_rows, 3):
        c1 = ws2.cell(row=r, column=1, value=lbl)
        c1.font = Font(bold=True, size=9, color=HULIOT_BLUE)
        c1.fill = PatternFill("solid", fgColor=HEADER_BG)
        c1.border = _border()
        c2 = ws2.cell(row=r, column=2, value=val)
        c2.font = Font(size=9)
        c2.border = _border()
        ws2.merge_cells(start_row=r, start_column=2,
                        end_row=r, end_column=4)

    r = 3 + len(info_rows) + 1
    _col_hdr(ws2, r, ["Section / Zone", "Total DU", "Flow (l/s)", "Req. Pipe Dia"])
    r += 1
    grand_du = 0
    for item in section_totals:
        for c, k in enumerate(["Section", "Total DU", "Flow (l/s)", "Req. Pipe Dia (mm)"], 1):
            cell = ws2.cell(row=r, column=c, value=item[k])
            cell.border = _border()
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        grand_du += item["Total DU"]
        r += 1

    # Grand total
    grand_flow = du_to_flow(grand_du)
    grand_dia, _ = flow_to_pipe(grand_flow)
    for c, v in enumerate(["GRAND TOTAL", round(grand_du,2),
                            grand_flow, f"{grand_dia} mm"], 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.font = Font(bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[r].height = 22

    # ── Sheet 3: Plumber Site Record ─────────────────────────────────────────
    ws3 = wb.create_sheet("Plumber Record")
    ws3.sheet_view.showGridLines = False
    for col_l, w in zip("ABCDE", [20, 20, 18, 18, 30]):
        ws3.column_dimensions[col_l].width = w

    ws3.merge_cells("A1:E1")
    t = ws3["A1"]
    t.value = "PLUMBER / SITE RECORD LOG"
    t.font = Font(bold=True, color=WHITE, size=12)
    t.fill = PatternFill("solid", fgColor=HULIOT_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26

    plumb_hdrs = ["Date", "Section", "Pipe Size (mm)", "Status", "Remarks"]
    _col_hdr(ws3, 3, plumb_hdrs)
    for r in range(4, 24):
        for c in range(1, 6):
            cell = ws3.cell(row=r, column=c, value="")
            cell.border = _border()
            cell.fill = PatternFill("solid",
                                    fgColor=ALT_ROW_BG if r % 2 == 0 else WHITE)
            ws3.row_dimensions[r].height = 20

    # ── Sheet 4: Builder BOQ ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Builder BOQ")
    ws4.sheet_view.showGridLines = False
    for col_l, w in zip("ABCDE", [30, 18, 12, 18, 22]):
        ws4.column_dimensions[col_l].width = w

    ws4.merge_cells("A1:E1")
    t4 = ws4["A1"]
    t4.value = "BILL OF QUANTITIES — DRAINAGE SYSTEM (HULIOT)"
    t4.font = Font(bold=True, color=WHITE, size=12)
    t4.fill = PatternFill("solid", fgColor=HULIOT_BLUE)
    t4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 26

    boq_hdrs = ["Description", "Pipe/Fitting Size", "Qty",
                 "Unit Rate (₹)", "Amount (₹)"]
    _col_hdr(ws4, 3, boq_hdrs)

    boq_items = []
    for item in section_totals:
        dia_str = item["Req. Pipe Dia (mm)"].replace(" mm", "")
        boq_items.append({
            "Description": f"Huliot HT Pro Pipe — {item['Section']}",
            "Pipe/Fitting Size": f"DN{dia_str}",
            "Qty": "",
            "Unit Rate (₹)": "",
            "Amount (₹)": f"=C{len(boq_items)+4}*D{len(boq_items)+4}",
        })

    for idx, item in enumerate(boq_items):
        r = 4 + idx
        bg = ALT_ROW_BG if idx % 2 == 0 else WHITE
        for c, k in enumerate(boq_hdrs, 1):
            key = k
            val = item.get(key, "")
            cell = ws4.cell(row=r, column=c, value=val)
            cell.border = _border()
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(size=9)
            cell.alignment = Alignment(
                horizontal="left" if c == 1 else "center",
                vertical="center")
        ws4.row_dimensions[r].height = 20

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════════════════════
# ── STREAMLIT UI ─────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

def main():
    # ── Header ────────────────────────────────────────────────────────────────
    col_logo, col_title = st.columns([1, 8])
    with col_logo:
        st.markdown("## 🔧")
    with col_title:
        st.markdown('<p class="main-title">Huliot DU Calculation Web App</p>', unsafe_allow_html=True)
        st.markdown('<p class="sub-title">EN 12056-2 · Upload → Edit → Download · No installation needed</p>', unsafe_allow_html=True)

    st.markdown("---")

    # ── Sidebar: Project Info & Settings ─────────────────────────────────────
    with st.sidebar:
        st.markdown("### 📋 Project Info")
        project_name = st.text_input("Project Name", placeholder="e.g. Lodha Palava T4")
        engineer     = st.text_input("Prepared By",  placeholder="e.g. Umesh Patil")
        date_val     = st.date_input("Date", value=datetime.today())

        st.markdown("---")
        st.markdown("### ⚙️ EN 12056-2 Settings")
        k_factor = st.selectbox(
            "K Factor (Discharge Frequency)",
            options=[0.5, 0.7, 1.0],
            index=1,
            help="0.5=Intermittent, 0.7=Frequent, 1.0=Congested"
        )
        global K_FACTOR
        K_FACTOR = k_factor

        st.markdown("---")
        st.markdown("### ℹ️ How to Use")
        st.markdown("""
1. **Upload** your DU Excel sheet
2. **Edit** QTY and Floors in the table
3. Results **auto-calculate**
4. **Download** formatted Excel
        """)
        st.markdown("---")
        st.caption("Huliot Pipes & Fittings Pvt. Ltd.\nWest Zone Technical Team")

    # ── File Upload ───────────────────────────────────────────────────────────
    st.markdown('<div class="section-head">📂 Upload DU Calculation Sheet</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader(
        "Upload Excel (.xlsx)", type=["xlsx"],
        help="Upload Huliot DU_Calculation_sheet.xlsx or compatible format"
    )

    if uploaded is None:
        st.markdown("""
<div class="warn-box">
⬆️ Please upload your <b>DU_Calculation_sheet.xlsx</b> file above to begin.
<br><br>The app will automatically detect sections (Kitchen, Toilet, etc.), 
let you edit quantities and floors, recalculate DU/Flow/Pipe size, 
and export a professional formatted Excel with project report + BOQ.
</div>
""", unsafe_allow_html=True)

        # Show sample data as demo
        st.markdown('<div class="section-head">📊 Sample Preview (Demo Data)</div>', unsafe_allow_html=True)
        demo = pd.DataFrame({
            "FIXTURES": ["Kitchen Sink", "WM 6 KG", "WC 6L Cistern", "Washbasin", "Floor Trap"],
            "DU": [0.8, 0.8, 2.0, 0.5, 0.6],
            "QTY": [1, 1, 1, 1, 1],
            "FLOORS": [10, 10, 10, 10, 10],
            "TOTAL DU": [8, 8, 20, 5, 6],
            "TOTAL FLOW (l/s)": [2.26, 2.26, 3.13, 1.57, 1.72],
            "REQ DIA (mm)": ["110 mm", "110 mm", "110 mm", "110 mm", "110 mm"],
        })
        st.dataframe(demo, use_container_width=True)
        return

    # ── Parse uploaded file ───────────────────────────────────────────────────
    file_bytes = uploaded.read()
    try:
        sections = parse_du_sheet(file_bytes)
    except Exception as e:
        st.error(f"❌ Could not parse file: {e}")
        return

    if not sections:
        st.warning("No recognisable DU data found. Check file format.")
        return

    st.success(f"✅ File loaded: **{uploaded.name}** — {len(sections)} section(s) detected")

    # ── Edit & Recalculate per section ────────────────────────────────────────
    project_info = {
        "project":  project_name or "—",
        "engineer": engineer or "—",
        "date":     str(date_val),
    }

    edited_sections = {}
    all_summary = []

    for sec_name, df in sections.items():
        st.markdown(f'<div class="section-head">🏗 {sec_name.upper()}</div>', unsafe_allow_html=True)

        edit_cols = ["FIXTURES", "QTY", "DU", "FLOORS"]
        display_df = df[edit_cols].copy()

        # Use data_editor for inline editing (QTY and FLOORS editable)
        with st.container():
            edited = st.data_editor(
                display_df,
                key=f"editor_{sec_name}",
                use_container_width=True,
                column_config={
                    "FIXTURES": st.column_config.TextColumn("Fixture", disabled=True, width="large"),
                    "QTY":    st.column_config.NumberColumn("Qty", min_value=0, max_value=100, step=1),
                    "DU":     st.column_config.NumberColumn("DU Value", min_value=0.0, max_value=10.0, step=0.1, disabled=True),
                    "FLOORS": st.column_config.NumberColumn("Floors", min_value=0, max_value=100, step=1),
                },
                num_rows="fixed",
                hide_index=True,
            )

        # Merge edits back into full df
        full_df = df.copy()
        full_df["QTY"]    = edited["QTY"].values
        full_df["FLOORS"] = edited["FLOORS"].values
        full_df = recalculate(full_df)

        # Results table
        result_df = full_df[["FIXTURES", "QTY", "DU", "FLOORS",
                              "TOTAL DU", "TOTAL FLOW (l/s)", "REQ DIA (mm)"]].copy()
        st.dataframe(result_df, use_container_width=True, hide_index=True)

        # Section KPIs
        total_du   = full_df["TOTAL DU"].sum()
        total_flow = du_to_flow(total_du)
        req_dia, _ = flow_to_pipe(total_flow)

        kpi1, kpi2, kpi3 = st.columns(3)
        with kpi1:
            st.markdown(f"""
<div class="metric-card">
  <div class="metric-label">TOTAL DU</div>
  <div class="metric-value">{total_du:.2f}</div>
</div>""", unsafe_allow_html=True)
        with kpi2:
            st.markdown(f"""
<div class="metric-card">
  <div class="metric-label">FLOW (l/s)</div>
  <div class="metric-value">{total_flow:.3f}</div>
</div>""", unsafe_allow_html=True)
        with kpi3:
            st.markdown(f"""
<div class="metric-card">
  <div class="metric-label">REQ. PIPE DIA</div>
  <div class="metric-value">{req_dia} mm</div>
</div>""", unsafe_allow_html=True)

        edited_sections[sec_name] = full_df
        all_summary.append({
            "Section": sec_name,
            "Total DU": round(total_du, 2),
            "Flow (l/s)": total_flow,
            "Pipe Dia (mm)": req_dia,
        })

    # ── Grand Summary ─────────────────────────────────────────────────────────
    st.markdown('<div class="section-head">📊 Grand Project Summary</div>', unsafe_allow_html=True)
    summary_df = pd.DataFrame(all_summary)
    st.dataframe(summary_df, use_container_width=True, hide_index=True)

    grand_du   = summary_df["Total DU"].sum()
    grand_flow = du_to_flow(grand_du)
    grand_dia, _ = flow_to_pipe(grand_flow)

    g1, g2, g3 = st.columns(3)
    with g1:
        st.metric("🏗 Grand Total DU", f"{grand_du:.2f}")
    with g2:
        st.metric("💧 Grand Total Flow", f"{grand_flow:.3f} l/s")
    with g3:
        st.metric("📏 Min. Main Stack Dia", f"{grand_dia} mm")

    # ── Download ──────────────────────────────────────────────────────────────
    st.markdown('<div class="section-head">⬇️ Download Results</div>', unsafe_allow_html=True)

    with st.spinner("Preparing formatted Excel…"):
        output_bytes = build_excel(edited_sections, project_info)

    fname = f"DU_Report_{project_name.replace(' ','_') or 'Project'}_{date_val}.xlsx"

    st.download_button(
        label="⬇️  Download Formatted DU Report (.xlsx)",
        data=output_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("""
<div class="warn-box">
📁 Downloaded file includes:<br>
&nbsp;&nbsp;• <b>DU Summary</b> — colour-coded fixture data per section<br>
&nbsp;&nbsp;• <b>Project Report</b> — grand totals and pipe recommendations<br>
&nbsp;&nbsp;• <b>Plumber Record</b> — blank site log for field team<br>
&nbsp;&nbsp;• <b>Builder BOQ</b> — pipe sizes per zone for procurement
</div>
""", unsafe_allow_html=True)


if __name__ == "__main__":
    main()
